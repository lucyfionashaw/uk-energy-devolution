import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

D=json.load(open('phase2_refresh.json'))
NAVY="1F3864"; BLUE="2E5496"; AMBER="FFE699"; GREEN="C6E0B4"; FNT="Arial"
wb=load_workbook('REPD_planning_outcomes_phase1.xlsx')

# remove sheets to refresh
for nm in ["Council control method","Outcomes by council party","Onshore wind by council party","Contested vs uncontested","Capacity vs count approval"]:
    if nm in wb.sheetnames: del wb[nm]

def hdr(ws,row,n,fill=BLUE):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name=FNT,bold=True,color="FFFFFF",size=10); cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def title(ws,t,s=None):
    ws["A1"]=t; ws["A1"].font=Font(name=FNT,bold=True,size=14,color=NAVY)
    if s: ws["A2"]=s; ws["A2"].font=Font(name=FNT,italic=True,size=9,color="595959")
def setw(ws,w):
    for k,v in w.items(): ws.column_dimensions[k].width=v
ORDER=["Con","Lab","LibDem","SNP","Plaid","Green","Reform","Sinn Fein","DUP","Other/Ind","Nationalist (pre-2007)"]
def ptable(ws,r0,data,label,capcols=True):
    cols=[label,"Projects","Granted","Refused","Approval rate"]+(["Capacity consented (MW)","Built (MW)"] if capcols else [])
    for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
    hdr(ws,r0,len(cols)); rr=r0+1
    for p in ORDER:
        if p not in data: continue
        o=data[p]
        ws.cell(row=rr,column=1,value=p); ws.cell(row=rr,column=2,value=o["n"]); ws.cell(row=rr,column=3,value=o.get("granted","")); ws.cell(row=rr,column=4,value=o.get("refused",""))
        ws.cell(row=rr,column=5,value=f"=IF((C{rr}+D{rr})=0,\"\",C{rr}/(C{rr}+D{rr}))"); ws.cell(row=rr,column=5).number_format="0.0%"
        if capcols:
            ws.cell(row=rr,column=6,value=o.get("grantedCap","")); ws.cell(row=rr,column=7,value=o.get("builtCap",""))
        for c in [2,3,4,6,7]:
            v=ws.cell(row=rr,column=c).value
            if isinstance(v,(int,float)): ws.cell(row=rr,column=c).number_format="#,##0"
        isni=p in ("Sinn Fein","DUP")
        ws.cell(row=rr,column=1).font=Font(name=FNT,size=10,bold=(p in("Con","Lab")),italic=isni)
        rr+=1
    return rr-1

# METHOD
ws=wb.create_sheet("Council control method"); ws.sheet_view.showGridLines=False
title(ws,"Phase 2 - matching projects to council political control","Who controlled the council when each decision was made. GB + Northern Ireland.")
cov=D["coverage"]
notes=[
 ("Council control source","Open Council Data UK (CC0): GB Compositions 1973-2026 (largest party by seats each year) + NI Compositions 2014-2026."),
 ("National government","Date lookup: Con (to May 1997), Lab (1997-2010), Con-led Coalition (2010-15), Con (2015-Jul 2024), Lab (Jul 2024-)."),
 ("Match date","Controlling party taken at the DECISION year (grant/refusal)."),
 ("",""),
 ("COVERAGE",""),
 ("Total project records",cov["total"]),
 ("National/strategic route (no council)",cov["nat"]),
 ("Local-route projects",cov["localRoute"]),
 ("Matched to a council",cov["matched"]),
 ("  of which Northern Ireland",cov["niMatched"]),
 ("Unmatched (edge cases only)",cov["unmatched"]),
 ("Match rate (GB+NI local route)","99.8% (13,370 of 13,397)"),
 ("",""),
 ("CAVEATS",""),
 ("Northern Ireland = largest party","NI councils use STV power-sharing and are almost always 'no overall control'. Shown as largest party (Sinn Fein or DUP), constant 2014-2026; pre-2014 NI decisions unmatched. Treat NI as a distinct system, not part of the Con/Lab story."),
 ("Largest-party proxy","GB control = largest party; coalitions/NOC attributed to the largest single party."),
 ("Pre-2007 'Nationalist'","Older Scottish/Welsh records combine SNP+Plaid as 'nat'; kept separate from modern SNP/Plaid."),
 ("Remaining 27 unmatched","Blanks, Isle of Man (not UK), 'NI Planning Service', and a few name typos/2024 renames - immaterial."),
]
r=4
for k,v in notes:
    ws.cell(row=r,column=1,value=k).font=Font(name=FNT,bold=True,size=10,color=NAVY)
    c=ws.cell(row=r,column=2,value=v); c.font=Font(name=FNT,size=10); c.alignment=Alignment(wrap_text=True,vertical="top")
    if isinstance(v,int): ws.cell(row=r,column=2).number_format="#,##0"
    r+=1
setw(ws,{"A":36,"B":95})

# OUTCOMES BY COUNCIL PARTY
ws=wb.create_sheet("Outcomes by council party"); ws.sheet_view.showGridLines=False
title(ws,"Planning outcomes by controlling council party","All technologies, control at decision year. NI councils (italic) shown as largest party.")
last=ptable(ws,4,D["byParty"],"Council controlling party")
ws.cell(row=last+2,column=1,value="Across all technologies the gap is modest (Con 89% - SNP 93%): in aggregate, council control is a weak predictor of approval. The signal is technology-specific (see onshore wind).").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":24,"B":10,"C":10,"D":10,"E":13,"F":22,"G":12})
ch=BarChart(); ch.type="bar"; ch.title="Approval rate by council party (all tech)"; ch.height=9; ch.width=15; ch.y_axis.numFmt="0%"
cats=Reference(ws,min_col=1,min_row=5,max_row=last); data=Reference(ws,min_col=5,min_row=4,max_row=last)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.legend=None
ws.add_chart(ch,"I4")

# ONSHORE BY PARTY
ws=wb.create_sheet("Onshore wind by council party"); ws.sheet_view.showGridLines=False
title(ws,"Onshore wind: approval by controlling council party","The contested technology - where local political control bites.")
last=ptable(ws,4,D["byPartyOnshore"],"Council controlling party")
for r in range(5,last+1):
    if ws.cell(r,1).value=="Con":
        for c in range(1,8): ws.cell(r,c).fill=PatternFill("solid",fgColor=AMBER)
ws.cell(row=last+2,column=1,value="GB: Conservative councils approve onshore wind least (61.5%) vs Lab 68.6%, LibDem 70.2%, SNP 72.9%. (NI councils approve ~88-90%, but NI onshore wind is a distinct, subsidy-driven story.)").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":24,"B":10,"C":10,"D":10,"E":13,"F":22,"G":12})
ch=BarChart(); ch.type="bar"; ch.title="Onshore wind approval rate by council party"; ch.height=9; ch.width=15; ch.y_axis.numFmt="0%"
cats=Reference(ws,min_col=1,min_row=5,max_row=last); data=Reference(ws,min_col=5,min_row=4,max_row=last)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.legend=None
ws.add_chart(ch,"I4")

# CONTESTED VS UNCONTESTED
ws=wb.create_sheet("Contested vs uncontested"); ws.sheet_view.showGridLines=False
title(ws,"Does council party matter? Onshore wind vs solar vs battery","Approval rate by controlling party, three technologies (GB parties).")
cols=["Council party","Onshore wind","Solar PV","Battery"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
hdr(ws,r0,len(cols)); rr=r0+1
for p in ["Con","Lab","LibDem","SNP","Other/Ind"]:
    ws.cell(row=rr,column=1,value=p)
    for ci,dset in enumerate([D["byPartyOnshore"],D["byPartySolar"],D["byPartyBattery"]],2):
        o=dset.get(p); ws.cell(row=rr,column=ci,value=(o["approvalRate"]/100 if o else None)); ws.cell(row=rr,column=ci).number_format="0.0%"
    ws.cell(row=rr,column=1).font=Font(name=FNT,size=10); rr+=1
last=rr-1
ws.cell(row=last+2,column=1,value="Spread across parties: onshore wind ~11 pts (61-73%); solar ~5 pts; battery ~6 pts. Partisanship shows up only on the contested technology.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":16,"B":14,"C":12,"D":12})
ch=BarChart(); ch.type="col"; ch.title="Approval rate by council party and technology"; ch.height=9; ch.width=16; ch.y_axis.numFmt="0%"
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last)
for col in [2,3,4]:
    data=Reference(ws,min_col=col,min_row=r0,max_row=last); ch.add_data(data,titles_from_data=True)
ch.set_categories(cats); ws.add_chart(ch,"F4")

# CAPACITY VS COUNT APPROVAL
ws=wb.create_sheet("Capacity vs count approval"); ws.sheet_view.showGridLines=False
title(ws,"Approval rate: by project count vs by capacity","Does weighting by MW change the approval rate? (Short answer: barely - unlike timing.)")
L=D["capLadder"]
ws.cell(row=4,column=1,value="Headline reconciliation").font=Font(name=FNT,bold=True,size=11,color=NAVY)
rows=[("Approval rate, by PROJECT COUNT  [granted / (granted+refused)]",L["countApproval"]/100,GREEN),
      ("Approval rate, by CAPACITY (MW)  [granted / (granted+refused)]",L["capApproval_grantedRefused"]/100,GREEN)]
r=5
for lab,val,fill in rows:
    ws.cell(row=r,column=1,value=lab); ws.cell(row=r,column=2,value=val); ws.cell(row=r,column=2).number_format="0.0%"
    for c in (1,2): ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor=fill)
    ws.cell(row=r,column=1).font=Font(name=FNT,size=10,bold=True); r+=1
ws.cell(row=r+1,column=1,value="Capacity-weighting does NOT lower the approval rate. Refused capacity is small (23 GW of 375 GW) and large projects are approved at similar or higher rates (offshore wind 96.6%). This is the opposite of timing, where capacity-weighting roughly quadruples the median (see 'Volume vs application wtd').").font=Font(name=FNT,italic=True,size=9,color="595959"); ws.cell(row=r+1,column=1).alignment=Alignment(wrap_text=True)
ws.merge_cells(start_row=r+1,start_column=1,end_row=r+2,end_column=7)

# ladder
lr=r+4
ws.cell(row=lr-1,column=1,value="Where lower numbers come from - it depends entirely on the definition (capacity-weighted)").font=Font(name=FNT,bold=True,size=11,color=NAVY)
ws.cell(row=lr,column=1,value="Definition (capacity-weighted)"); ws.cell(row=lr,column=2,value="Value")
hdr(ws,lr,2)
ladder=[("Granted / (granted + refused)  -- 'approval rate'",L["capApproval_grantedRefused"]),
 ("Granted / (granted + refused + withdrawn + abandoned)",L["capApproval_inclWithdrawnAbandoned"]),
 ("Granted / (decided + still-pending)",L["capApproval_inclPending"]),
 ("Consented / total capacity ever in REPD",L["consentedOverTotal"]),
 ("Build-out: (operational + under construction) / consented",L["buildOut_opUC_overConsented"]),
 ("Delivered: (operational + under construction) / total",L["delivered_opUC_overTotal"]),
 ("Operational / total",L["operationalOverTotal"])]
rr=lr+1
for lab,val in ladder:
    ws.cell(row=rr,column=1,value=lab); ws.cell(row=rr,column=2,value=val/100); ws.cell(row=rr,column=2).number_format="0.0%"; ws.cell(row=rr,column=1).font=Font(name=FNT,size=10)
    if 38<=val<=46:
        for c in (1,2): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=AMBER)
    rr+=1
ws.cell(row=rr+1,column=1,value="A 'just 40-44%' figure does not match any APPROVAL definition here. It is in the range of DELIVERY / build-out metrics (consents that actually get built: 35%) or a recent-applications cohort where much capacity is still pending. If your prior figure was an approval rate, it likely used a different REPD vintage or population - worth pinning down the exact denominator.").font=Font(name=FNT,italic=True,size=9,color="C00000"); ws.cell(row=rr+1,column=1).alignment=Alignment(wrap_text=True)
ws.merge_cells(start_row=rr+1,start_column=1,end_row=rr+3,end_column=7)

# capacity by status (GW)
gr=rr+5
ws.cell(row=gr-1,column=1,value="Capacity by status (GW) - the building blocks").font=Font(name=FNT,bold=True,size=11,color=NAVY)
ws.cell(row=gr,column=1,value="Bucket"); ws.cell(row=gr,column=2,value="Capacity (GW)")
hdr(ws,gr,2)
kc=D["keyCapsGW"]; rr=gr+1
for lab,key in [("Consented (granted, incl. awaiting/built/op)","consented"),("  of which Awaiting Construction (unbuilt consents)","awaitingConstruction"),("  of which Under Construction","underConstruction"),("  of which Operational","operational"),("Refused","refused"),("Withdrawn","withdrawn"),("Abandoned","abandoned"),("Still pending a decision","pending"),("TOTAL","total")]:
    ws.cell(row=rr,column=1,value=lab); ws.cell(row=rr,column=2,value=kc[key]); ws.cell(row=rr,column=2).number_format="#,##0.0"
    ws.cell(row=rr,column=1).font=Font(name=FNT,size=10,bold=(key in("consented","total")))
    rr+=1
setw(ws,{"A":52,"B":14,"C":4,"D":12,"E":12,"F":12,"G":12})

order=['Read me','Overview','Funnel by technology','Funnel by nation','Durations','Time to decision','Volume vs application wtd','Capacity vs count approval','Decided vs awaiting','Solar timing explained','Decision route','Approval trend','Onshore wind by nation','Applications by tech','Decommissioned','Council control method','Outcomes by council party','Onshore wind by council party','Contested vs uncontested','Council vs national','MP party method','Outcomes by MP party','Council vs MP']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save('REPD_planning_outcomes_phase1.xlsx')
print("saved",len(wb.sheetnames),"sheets")
