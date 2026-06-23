import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

D=json.load(open('phase2b_data.json'))
NAVY="1F3864"; BLUE="2E5496"; AMBER="FFE699"; GREEN="C6E0B4"; FNT="Arial"
wb=load_workbook('REPD_planning_outcomes_phase1.xlsx')

def hdr(ws,row,n,fill=BLUE):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name=FNT,bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def title(ws,t,s=None):
    ws["A1"]=t; ws["A1"].font=Font(name=FNT,bold=True,size=14,color=NAVY)
    if s: ws["A2"]=s; ws["A2"].font=Font(name=FNT,italic=True,size=9,color="595959")
def setw(ws,w):
    for k,v in w.items(): ws.column_dimensions[k].width=v
ORDER=["Con","Lab","LibDem","SNP","Plaid","Green","Reform","Other"]
def ptable(ws,r0,data,label):
    cols=[label,"Projects","Granted","Refused","Approval rate","Capacity consented (MW)"]
    for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
    hdr(ws,r0,len(cols)); rr=r0+1
    for p in ORDER:
        if p not in data: continue
        o=data[p]
        ws.cell(row=rr,column=1,value=p+" MP"); ws.cell(row=rr,column=2,value=o["n"]); ws.cell(row=rr,column=3,value=o["granted"]); ws.cell(row=rr,column=4,value=o["refused"])
        ws.cell(row=rr,column=5,value=f"=IF((C{rr}+D{rr})=0,\"\",C{rr}/(C{rr}+D{rr}))"); ws.cell(row=rr,column=5).number_format="0.0%"
        ws.cell(row=rr,column=6,value=o.get("grantedCap",""))
        for c in [2,3,4,6]: ws.cell(row=rr,column=c).number_format="#,##0"
        ws.cell(row=rr,column=1).font=Font(name=FNT,size=10,bold=(p in("Con","Lab")))
        rr+=1
    return rr-1

# METHOD
ws=wb.create_sheet("MP party method"); ws.sheet_view.showGridLines=False
title(ws,"Phase 2b - MP / constituency party at decision date","Each project geocoded to its Westminster constituency; MP party = winner of the general election in effect at the decision.")
cov=D["coverage"]
notes=[
 ("Geocoding","Project postcode -> Westminster constituency via postcodes.io (ONS-derived). Both 2010-2024 and 2024 boundary constituencies returned."),
 ("MP party source","House of Commons Library: constituency results 1918-2019 (CBP-8647) for GE2010/2015/2017/2019, and GE2024 results (CBP-10009). Winner = largest vote share."),
 ("Match date","GE in effect at the decision date: 2010 boundaries/results for decisions May 2010-Jul 2024; 2024 boundaries/results from Jul 2024."),
 ("",""),
 ("COVERAGE",""),
 ("Total project records",cov["total"]),
 ("National/strategic route (excluded)",cov["nat"]),
 ("Geocoded to a constituency",cov["geocoded"]),
 ("Matched to an MP party at decision",cov["mpMatched"]),
 ("Have BOTH council and MP party",cov["councilAndMp"]),
 ("",""),
 ("IMPORTANT - how to read this",""),
 ("MP party is an AREA proxy, not the decision-maker","Planning decisions are made by COUNCILS, not MPs. MP party mainly indicates the TYPE of area (rural Tory shire, urban Labour, etc.). So the MP cut describes where projects sit, not who approved them."),
 ("The apparent reversal","Conservative-MP areas show the HIGHEST onshore-wind approval (rural areas host most wind and much was consented, incl. Scotland's national route), even though Conservative-CONTROLLED COUNCILS are the most restrictive. Do not conflate the two - this is an ecological-inference trap."),
 ("Smaller, noisier sample","The MP layer (5,805) is roughly half the council layer (12,805) due to missing/old postcodes and 2010<->2024 boundary renames. Onshore-wind MP cells are small (n=26-115); treat as indicative."),
]
r=4
for k,v in notes:
    ws.cell(row=r,column=1,value=k).font=Font(name=FNT,bold=True,size=10,color=NAVY)
    c=ws.cell(row=r,column=2,value=v); c.font=Font(name=FNT,size=10); c.alignment=Alignment(wrap_text=True,vertical="top")
    if isinstance(v,int): ws.cell(row=r,column=2).number_format="#,##0"
    r+=1
setw(ws,{"A":34,"B":95})

# OUTCOMES BY MP PARTY
ws=wb.create_sheet("Outcomes by MP party"); ws.sheet_view.showGridLines=False
title(ws,"Planning outcomes by constituency MP party","MP party at the decision date. NB: an area proxy, not the decision-maker (councils decide).")
ws.cell(row=4,column=1,value="All technologies").font=Font(name=FNT,bold=True,size=11,color=NAVY)
last=ptable(ws,5,D["byMP"],"Constituency MP party")
r2=last+3
ws.cell(row=r2-1,column=1,value="Onshore wind only (small samples)").font=Font(name=FNT,bold=True,size=11,color=NAVY)
last2=ptable(ws,r2,D["byMPon"],"Constituency MP party")
for r in range(r2+1,last2+1):
    if ws.cell(r,1).value=="Con MP":
        for c in range(1,7): ws.cell(r,c).fill=PatternFill("solid",fgColor=AMBER)
ws.cell(row=last2+2,column=1,value="Note the reversal vs council control: Con-MP AREAS approve onshore wind most (rural siting), while Con-CONTROLLED COUNCILS approve it least. MP party = area type, not decision-maker.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":24,"B":11,"C":11,"D":11,"E":14,"F":22})
ch=BarChart(); ch.type="bar"; ch.title="Onshore wind approval by constituency MP party"; ch.height=8; ch.width=14; ch.y_axis.numFmt="0%"
cats=Reference(ws,min_col=1,min_row=r2+1,max_row=last2); data=Reference(ws,min_col=5,min_row=r2,max_row=last2)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.legend=None
ws.add_chart(ch,"H"+str(r2-1))

# DIVERGENCE
ws=wb.create_sheet("Council vs MP"); ws.sheet_view.showGridLines=False
title(ws,"When council party and MP party differ","Does it matter whether the controlling council and the local MP are the same party?")
def drows(ws,r0,data,title_txt):
    ws.cell(row=r0-1,column=1,value=title_txt).font=Font(name=FNT,bold=True,size=11,color=NAVY)
    cols=["","Projects","Granted","Refused","Approval rate"]
    for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
    hdr(ws,r0,5); rr=r0+1
    for lab,key in [("Council and MP same party","same"),("Council and MP different parties","diff")]:
        o=data[key];ws.cell(row=rr,column=1,value=lab);ws.cell(row=rr,column=2,value=o["n"]);ws.cell(row=rr,column=3,value=o["granted"]);ws.cell(row=rr,column=4,value=o["refused"])
        ws.cell(row=rr,column=5,value=f"=C{rr}/(C{rr}+D{rr})");ws.cell(row=rr,column=5).number_format="0.0%"
        for c in [2,3,4]: ws.cell(row=rr,column=c).number_format="#,##0"
        rr+=1
    return rr-1
l1=drows(ws,5,D["divAll"],"All technologies")
l2=drows(ws,l1+3,D["divOn"],"Onshore wind only")
# combos
cr=l2+3
ws.cell(row=cr-1,column=1,value="Onshore wind: approval by council party / MP party combination (n>=15)").font=Font(name=FNT,bold=True,size=11,color=NAVY)
cols=["Council / MP combination","Projects","Granted","Refused","Approval rate"]
for i,h in enumerate(cols,1): ws.cell(row=cr,column=i,value=h)
hdr(ws,cr,5); rr=cr+1
for k,o in D["onComboTop"]:
    ws.cell(row=rr,column=1,value=k);ws.cell(row=rr,column=2,value=o["n"]);ws.cell(row=rr,column=3,value=o["granted"]);ws.cell(row=rr,column=4,value=o["refused"])
    ws.cell(row=rr,column=5,value=f"=C{rr}/(C{rr}+D{rr})");ws.cell(row=rr,column=5).number_format="0.0%"
    for c in [2,3,4]: ws.cell(row=rr,column=c).number_format="#,##0"
    ws.cell(row=rr,column=1).font=Font(name=FNT,size=10)
    rr+=1
ws.cell(row=rr+1,column=1,value="Alignment between council and MP makes little difference (all-tech 92.7% vs 91.5%; onshore 76.7% vs 74.0%). The operative actor is the council.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":34,"B":11,"C":11,"D":11,"E":14})

order=['Read me','Overview','Funnel by technology','Funnel by nation','Durations','Time to decision','Volume vs application wtd','Decided vs awaiting','Solar timing explained','Decision route','Approval trend','Onshore wind by nation','Applications by tech','Decommissioned','Council control method','Outcomes by council party','Onshore wind by council party','Contested vs uncontested','Council vs national','MP party method','Outcomes by MP party','Council vs MP']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save('REPD_planning_outcomes_phase1.xlsx')
print("saved",len(wb.sheetnames),"sheets")
