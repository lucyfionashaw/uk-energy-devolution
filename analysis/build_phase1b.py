import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

D=json.load(open('phase1b_data.json'))
NAVY="1F3864"; BLUE="2E5496"; GREY="F2F2F2"; AMBER="FFE699"; GREEN="C6E0B4"; FNT="Arial"
wb=load_workbook('REPD_planning_outcomes_phase1.xlsx')

def hdr(ws,row,n,fill=BLUE):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name=FNT,bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def title(ws,t,s=None):
    ws["A1"]=t; ws["A1"].font=Font(name=FNT,bold=True,size=14,color=NAVY)
    if s: ws["A2"]=s; ws["A2"].font=Font(name=FNT,italic=True,size=9,color="595959")
def setw(ws,w):
    for k,v in w.items(): ws.column_dimensions[k].width=v

# ---- TIME TO DECISION ----
ws=wb.create_sheet("Time to decision"); ws.sheet_view.showGridLines=False
title(ws,"Time to a planning decision — approvals AND refusals","Median months, application to first decision. Refusals also consume time; survivor-based.")
cols=["Technology","Granted: n","Granted: median months","Refused: n","Refused: median months"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
hdr(ws,r0,len(cols))
rr=r0+1
for row in D["decTbl"]:
    ws.cell(row=rr,column=1,value=row["tech"]); ws.cell(row=rr,column=2,value=row["nGrant"])
    ws.cell(row=rr,column=3,value=row["medGrant"]); ws.cell(row=rr,column=4,value=row["nRefuse"]); ws.cell(row=rr,column=5,value=row["medRefuse"])
    for c in [3,5]:
        if ws.cell(row=rr,column=c).value is not None: ws.cell(row=rr,column=c).number_format="0.0"
    for c in [2,4]: ws.cell(row=rr,column=c).number_format="#,##0"
    ws.cell(row=rr,column=1).font=Font(name=FNT,size=10)
    if row["tech"] in ("Wind Onshore","Wind Offshore","Solar Photovoltaics"):
        for c in range(1,6): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=GREY)
    rr+=1
last=rr-1
ws.cell(row=last+2,column=1,value="Onshore wind takes a median 12.4 months even to be REFUSED (882 cases) — substantial runway spent on projects that fail.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":32,"B":12,"C":22,"D":12,"E":22})
ch=BarChart(); ch.type="bar"; ch.title="Median months to decision: granted vs refused"; ch.height=10; ch.width=17
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last)
for col in [3,5]:
    data=Reference(ws,min_col=col,min_row=r0,max_row=last); ch.add_data(data,titles_from_data=True)
ch.set_categories(cats)
ws.add_chart(ch,"G4")

# decided vs undecided + reasonableness
ws=wb.create_sheet("Decided vs awaiting"); ws.sheet_view.showGridLines=False
title(ws,"Decided vs still-awaiting a decision","Of 14,316 records, those with a recorded grant/refusal date vs none. Pending split by 'reasonableness' of the wait.")
ws.cell(row=4,column=1,value="Has a recorded planning decision?").font=Font(name=FNT,bold=True,size=11,color=NAVY)
du=D["decidedUndecided"]
ws.cell(row=5,column=1,value="With grant/refusal date"); ws.cell(row=5,column=2,value=du["decided"])
ws.cell(row=6,column=1,value="No decision date yet"); ws.cell(row=6,column=2,value=du["undecided"])
ws.cell(row=7,column=1,value="Total"); ws.cell(row=7,column=2,value="=B5+B6")
for r in (5,6,7):
    ws.cell(row=r,column=2).number_format="#,##0"; ws.cell(row=r,column=1).font=Font(name=FNT,size=10,bold=(r==7))
ws.cell(row=5,column=2).fill=PatternFill("solid",fgColor=GREEN); ws.cell(row=6,column=2).fill=PatternFill("solid",fgColor=AMBER)
# breakdown of undecided by status
ws.cell(row=9,column=1,value="Where 'no decision date' records sit (current status)").font=Font(name=FNT,bold=True,size=11,color=NAVY)
ws.cell(row=10,column=1,value="Current status"); ws.cell(row=10,column=2,value="Records")
hdr(ws,10,2)
rr=11
for st,n in sorted(du["undecByStatus"].items(),key=lambda x:-x[1]):
    ws.cell(row=rr,column=1,value=st); ws.cell(row=rr,column=2,value=n); ws.cell(row=rr,column=2).number_format="#,##0"; ws.cell(row=rr,column=1).font=Font(name=FNT,size=10)
    rr+=1
note=rr+1
ws.cell(row=note,column=1,value="Note: 'Operational'/'Under Construction' with no grant date are data gaps, not genuinely undecided. Genuinely awaiting = 'Application Submitted' (1,673) + 'Appeal Lodged'.").font=Font(name=FNT,italic=True,size=9,color="595959")
ws.cell(row=note+1,column=1,value="'Withdrawn'/'Abandoned'/'Revised' with no decision = projects that exited before any decision.").font=Font(name=FNT,italic=True,size=9,color="595959")

# reasonableness bands (right side)
ws.cell(row=4,column=4,value="Genuinely-pending projects: how long they have waited").font=Font(name=FNT,bold=True,size=11,color=NAVY)
ws.cell(row=5,column=4,value="Wait since submission"); ws.cell(row=5,column=5,value="Projects"); ws.cell(row=5,column=6,value="Capacity (MW)")
hdr(ws,5,6);
# only style D5:F5
for c in (4,5,6):
    cell=ws.cell(row=5,column=c); cell.font=Font(name=FNT,bold=True,color="FFFFFF",size=10); cell.fill=PatternFill("solid",fgColor=BLUE); cell.alignment=Alignment(horizontal="center",wrap_text=True)
order=["<4 mo (within statutory)","4-12 mo","1-2 yrs","2-3 yrs","3-5 yrs","5+ yrs"]
pb=D["pendingBands"]; rr=6
for b in order:
    ws.cell(row=rr,column=4,value=b); ws.cell(row=rr,column=5,value=pb[b]["n"]); ws.cell(row=rr,column=6,value=pb[b]["cap"])
    ws.cell(row=rr,column=5).number_format="#,##0"; ws.cell(row=rr,column=6).number_format="#,##0"; ws.cell(row=rr,column=4).font=Font(name=FNT,size=10)
    if b=="<4 mo (within statutory)":
        for c in (4,5,6): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=GREEN)
    if b in ("2-3 yrs","3-5 yrs","5+ yrs"):
        for c in (4,5,6): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=AMBER)
    rr+=1
ws.cell(row=rr+1,column=4,value="Statutory determination: ~8 weeks (minor), ~13-16 weeks (major/EIA). Only ~308 projects are within that window; ~878 (>1yr) are genuinely overdue.").font=Font(name=FNT,italic=True,size=9,color="595959")
ws.column_dimensions["D"].width=26
setw(ws,{"A":30,"B":12,"C":4,"D":26,"E":12,"F":14})

# ---- SOLAR TIMING ----
ws=wb.create_sheet("Solar timing explained"); ws.sheet_view.showGridLines=False
title(ws,"Why solar's median is just 2.3 months","It is genuine application-to-permission time, but a size-mix effect: most solar is small and fast.")
sd=D["solarDiag"]
ws.cell(row=4,column=1,value="Median app->grant by project size").font=Font(name=FNT,bold=True,size=11,color=NAVY)
ws.cell(row=5,column=1,value="Capacity band"); ws.cell(row=5,column=2,value="Projects"); ws.cell(row=5,column=3,value="Median months")
hdr(ws,5,3)
rr=6
for b in ["<1MW","1-5MW","5-50MW","50MW+"]:
    o=sd["byCap"][b]; ws.cell(row=rr,column=1,value=b); ws.cell(row=rr,column=2,value=o["n"]); ws.cell(row=rr,column=3,value=o["med"])
    ws.cell(row=rr,column=2).number_format="#,##0"; ws.cell(row=rr,column=3).number_format="0.0"; ws.cell(row=rr,column=1).font=Font(name=FNT,size=10)
    rr+=1
ws.cell(row=rr+1,column=1,value=f"{sd['within3']}% of solar decided within 3 months; {sd['within6']}% within 6; only {sd['pctZeroToOne']}% in under 1 month (so not a same-date data artifact).").font=Font(name=FNT,italic=True,size=9,color="595959")
ws.cell(row=rr+2,column=1,value="Large 50MW+ solar (NSIP route) takes ~16 months — the same order as onshore wind. The 2.3mo headline is dominated by 3,000 sub-1MW schemes.").font=Font(name=FNT,italic=True,size=9,color="595959")
# sample pipeline
sr=rr+4
ws.cell(row=sr-1,column=1,value="Sample pipeline (illustrative records, small -> large)").font=Font(name=FNT,bold=True,size=11,color=NAVY)
cols=["Site","Planning authority","Capacity (MW)","Submitted","Granted","Months"]
for i,h in enumerate(cols,1): ws.cell(row=sr,column=i,value=h)
hdr(ws,sr,len(cols))
rr=sr+1
for s in D["solarSample"]:
    ws.cell(row=rr,column=1,value=s["site"]); ws.cell(row=rr,column=2,value=s["auth"]); ws.cell(row=rr,column=3,value=s["cap"])
    ws.cell(row=rr,column=4,value=s["sub"]); ws.cell(row=rr,column=5,value=s["grant"]); ws.cell(row=rr,column=6,value=s["m"])
    ws.cell(row=rr,column=3).number_format="#,##0.00"; ws.cell(row=rr,column=6).number_format="0.0"
    for c in range(1,7): ws.cell(row=rr,column=c).font=Font(name=FNT,size=10)
    if "NSIP" in s["auth"]:
        for c in range(1,7): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=AMBER)
    rr+=1
setw(ws,{"A":26,"B":30,"C":14,"D":12,"E":12,"F":10})
ch=BarChart(); ch.type="col"; ch.title="Median months to consent by solar project size"; ch.height=8; ch.width=12
cats=Reference(ws,min_col=1,min_row=6,max_row=9); data=Reference(ws,min_col=3,min_row=5,max_row=9)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.legend=None
ws.add_chart(ch,"H4")

# ---- update Read me data notes ----
rm=wb["Read me"]
# find next empty row
nr=rm.max_row+2
extra=[
 ("DATA-INTEGRITY NOTES (added)",""),
 ("Full file vs preview","The downloadable CSV is the COMPLETE database (14,316 rows). The gov.uk on-page preview shows only ~200 rows ('subset'); the file itself is not a subset."),
 ("Withdrawn rows retained","DESNZ does not delete withdrawn/refused/abandoned projects - all are present. Bias runs the other way: data is 'sourced mostly from web and developers', so dead projects may linger un-updated."),
 ("150kW threshold","Minimum capacity 1MW until 2021, then 150kW. Sub-1MW projects in planning before 2021 are largely absent - relevant to solar counts/timings."),
 ("Stuck = maybe stale","Some 'pending 5+ years' projects are likely abandoned-but-not-updated rather than genuinely awaiting a decision."),
 ("Manager","Barbour ABI manages REPD for DESNZ; snapshot updated the month after each quarter-end."),
]
for k,v in extra:
    rm.cell(row=nr,column=1,value=k).font=Font(name=FNT,bold=True,size=10,color=NAVY)
    c=rm.cell(row=nr,column=2,value=v); c.font=Font(name=FNT,size=10); c.alignment=Alignment(wrap_text=True,vertical="top")
    nr+=1

# reorder: move new analytical sheets after 'Durations'
order=['Read me','Overview','Funnel by technology','Funnel by nation','Durations','Time to decision','Decided vs awaiting','Solar timing explained','Decision route','Approval trend','Onshore wind by nation','Applications by tech','Decommissioned']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save('REPD_planning_outcomes_phase1.xlsx')
print("updated", wb.sheetnames)
