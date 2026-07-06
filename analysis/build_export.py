#!/usr/bin/env python3
"""Export every REPD project with the analysis appendages bolted on, so the joins can be
audited by hand: controlling council party at the decision date, that council's land area,
the local/national route, the resubmission/dedup flags, and the original application date.

Output: data/processed/repd_enriched.xlsx  (two sheets: Projects, Readme)
Run:    python3 analysis/build_export.py
"""
import csv, datetime, pathlib, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = pathlib.Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import council_control as CC
import repd_records as R
import geo_resolve as G

ROOT = HERE.parent
OUT = ROOT / "data" / "processed" / "repd_enriched.xlsx"

def num(x):
    x = (x or "").strip().replace(",", "")
    try: return float(x)
    except ValueError: return None
def has(x): return bool((x or "").strip())
def pdate(s):
    s = (s or "").strip()
    try:
        d, m, y = s.split("/"); return datetime.date(int(y), int(m), int(d))
    except (ValueError, IndexError): return None
def iso(d): return d.isoformat() if d else ""

GD = ["Planning Permission  Granted", "Appeal Granted", "Secretary of State - Granted"]
RD = ["Planning Permission Refused", "Appeal Refused", "Secretary of State - Refusal"]
GRANT_ST = {"Awaiting Construction", "Under Construction", "Operational", "Planning Permission Expired", "Decommissioned"}
REF_ST = {"Application Refused", "Appeal Refused", "Secretary of State Refusal"}

def months(a, b):
    if not a or not b: return None
    v = (b.year - a.year) * 12 + (b.month - a.month) + (b.day - a.day) / 30.0
    return round(v, 1) if 0 <= v < 600 else None

HEAD = ["Ref ID", "Site Name", "Technology", "Capacity MW", "Country", "Planning Authority",
        "Council party (at decision)", "Council land km²", "Route", "Council match note",
        "Status", "Superseded duplicate?", "Replaced by (Ref)",
        "Submitted (own)", "Original submitted", "Decision date", "Under construction", "Operational",
        "Months: original submission → decision"]

wb = Workbook()
ws = wb.active; ws.title = "Projects"
ws.append(HEAD)
for c in range(1, len(HEAD) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="2E5496")
    cell.alignment = Alignment(wrap_text=True, vertical="top")

for x in R.load():                      # every row, superseded flagged
    st = x["Development Status (short)"].strip()
    granted = has(x[GD[0]]) or st in GRANT_ST
    refused = st in REF_ST
    dec = None
    if granted: dec = next((pdate(x[c]) for c in GD if has(x[c])), None)
    elif refused: dec = next((pdate(x[c]) for c in RD if has(x[c])), None)
    orig = R.original_submission(x)
    # The named planning authority IS the authority. Control is read at the DECISION date; if
    # there is no decision date we cannot know who was in office, so the party is left blank.
    auth, route, note = G.resolve(x["Planning Authority"], x["Country"])
    if auth and dec:
        party = CC.control_at(auth, dec)
    elif auth and not dec:
        party = None
        if not note:
            note = "no decision date recorded — cannot identify the council in office"
    else:
        party = None
    km2 = G.land_km2(auth) if auth else None
    ws.append([
        x["Ref ID"].strip(), x["Site Name"].strip(), x["Technology Type"].strip(),
        num(x["Installed Capacity (MWelec)"]), x["Country"].strip(), x["Planning Authority"].strip(),
        party or "", km2 if km2 is not None else "", route, note,
        st, "Y" if x["_superseded"] else "", x["Are they re-applying (New REPD Ref)"].strip(),
        iso(x["_sub"]), iso(orig), iso(dec), iso(pdate(x["Under Construction"])), iso(pdate(x["Operational"])),
        months(orig, dec) if dec else "",
    ])
ws.freeze_panes = "A2"
widths = {"A": 10, "B": 34, "C": 20, "D": 11, "E": 10, "F": 26, "G": 20, "H": 13, "I": 15,
          "J": 46, "K": 22, "L": 15, "M": 16, "N": 14, "O": 16, "P": 14, "Q": 16, "R": 14,
          "S": 26}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

rm = wb.create_sheet("Readme")
notes = [
    ("REPD enriched export", ""),
    ("Source", "Renewable Energy Planning Database (REPD) Q1 2026, DESNZ."),
    ("Rows", "Every application. 'Superseded duplicate? = Y' marks the 815 resubmitted rows that "
             "are replaced by a newer record (dropped from the site's charts)."),
    ("Council party (at decision)", "Party controlling the NAMED planning authority on the DECISION "
             "date, from Open Council Data UK's `majority` control column — coalitions credited to the "
             "lead party. Control is read at the decision date honouring the May take-office cut-off "
             "(decisions before mid-May fall under the previous administration). Blank where no council "
             "applies or the decision date is missing (see 'Council match note')."),
    ("Council land km²", "ONS Standard Area Measurement land area of the named authority. The 11 "
             "unitaries created by the 2019-2023 reorganisation = sum of their former districts' "
             "areas. Northern Ireland areas are shown but excluded from the GB land-density chart. "
             "Blank for national-route / unmatched rows."),
    ("Route", "'Local council' = matched to a GB lower-tier council; 'National' = a national "
             "consenting body (Scottish Government S36, Planning Inspectorate NSIP, Crown Estate, "
             "Marine Scotland, NI Planning Service, DECC S36, Marine Management Organisation, Welsh "
             "Government NSIP) with no local council; 'Northern Ireland' = NI council (own system); "
             "'Outside GB' = Crown Dependency (Isle of Man, Jersey); 'Unmatched' = a name we could "
             "not resolve (typo, parliamentary constituency, place name or blank)."),
    ("Council match note", "Why a row has no council party: no decision date, national route, "
             "Northern Ireland, outside GB, or an unrecognised authority name (with the raw name)."),
    ("Original submitted", "Earliest application date across the resubmission chain (used for the true "
             "time-to-decide)."),
    ("Months: original → decision", "Months from the ORIGINAL application to the decision date."),
]
for i, (k, v) in enumerate(notes, 1):
    rm.cell(row=i, column=1, value=k).font = Font(bold=True)
    rm.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
rm.column_dimensions["A"].width = 30
rm.column_dimensions["B"].width = 100

wb.save(OUT)
print(f"wrote {OUT}  ({OUT.stat().st_size:,} bytes; {ws.max_row-1:,} project rows)")
