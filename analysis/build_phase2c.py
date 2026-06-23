import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

D=json.load(open('phase2c_data.json'))
NAVY="1F3864"; BLUE="2E5496"; AMBER="FFE699"; GREEN="C6E0B4"; FNT="Arial"
wb=load_workbook('REPD_planning_outcomes_phase1.xlsx')

def hdr(ws,row,n,fill=BLUE):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name=FNT,bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def title(ws,t,s=None):
    ws["A1"]=t; ws["A1"].font=Font(name=FNT,bold=True,size=14,color=NAVY)
    if s: ws["A2"]=s; ws["A2"].font=Font(name=FNT,italic=True,size=9,color="595959")
def setw(ws,w):
    for k,v in w.items(): ws.column_dimensions[k].width=v

ws=wb.create_sheet("Volume vs application wtd"); ws.sheet_view.showGridLines=False
title(ws,"Volume-weighted vs application-weighted timing","Median months, application to consent. Count-weighted treats every project equally; capacity-weighted weights by MW.")
cols=["Technology","Projects","By project count (months)","By capacity / MW (months)","Slowdown x"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
hdr(ws,r0,len(cols))
rr=r0+1
for tech,o in D["weighted"].items():
    ws.cell(row=rr,column=1,value=tech); ws.cell(row=rr,column=2,value=o["n"])
    ws.cell(row=rr,column=3,value=o["countMedian"]); ws.cell(row=rr,column=4,value=o["capWeightedMedian"])
    ws.cell(row=rr,column=5,value=f"=D{rr}/C{rr}")
    ws.cell(row=rr,column=2).number_format="#,##0"
    for c in [3,4]: ws.cell(row=rr,column=c).number_format="0.0"
    ws.cell(row=rr,column=5).number_format="0.0x"
    ws.cell(row=rr,column=1).font=Font(name=FNT,size=10,bold=(tech in("Solar PV","All technologies")))
    if tech in ("Solar PV","All technologies"):
        for c in range(1,6): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=AMBER)
    rr+=1
last=rr-1
ws.cell(row=last+2,column=1,value="Most PROJECTS are small and fast; most MEGAWATTS sit in large, slow schemes. Solar's 2.3-month headline becomes 9.7 months once weighted by capacity; all-tech 3.8 -> 14.4.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":20,"B":11,"C":22,"D":22,"E":12})
ch=BarChart(); ch.type="col"; ch.title="Median months to consent: by count vs by capacity"; ch.height=8; ch.width=15
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last)
for col in [3,4]:
    data=Reference(ws,min_col=col,min_row=r0,max_row=last); ch.add_data(data,titles_from_data=True)
ch.set_categories(cats)
ws.add_chart(ch,"G4")

# solar by size band
sr=last+5
ws.cell(row=sr-1,column=1,value="Solar: time to consent by project size (the 50MW NSIP cliff)").font=Font(name=FNT,bold=True,size=11,color=NAVY)
cols2=["Capacity band","Projects","Median months","Total capacity (MW)"]
for i,h in enumerate(cols2,1): ws.cell(row=sr,column=i,value=h)
hdr(ws,sr,len(cols2))
rr=sr+1
for b in D["solarBins"]:
    ws.cell(row=rr,column=1,value=b["band"]); ws.cell(row=rr,column=2,value=b["n"]); ws.cell(row=rr,column=3,value=b["medMonths"]); ws.cell(row=rr,column=4,value=b["capMW"])
    ws.cell(row=rr,column=2).number_format="#,##0"; ws.cell(row=rr,column=3).number_format="0.0"; ws.cell(row=rr,column=4).number_format="#,##0"
    ws.cell(row=rr,column=1).font=Font(name=FNT,size=10)
    if b["band"]=="49.5-50":
        for c in range(1,5): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=GREEN)
    rr+=1
slast=rr-1
ch2=BarChart(); ch2.type="col"; ch2.title="Solar: median months to consent by size band"; ch2.height=8; ch2.width=14; ch2.legend=None
cats=Reference(ws,min_col=1,min_row=sr+1,max_row=slast); data=Reference(ws,min_col=3,min_row=sr,max_row=slast)
ch2.add_data(data,titles_from_data=True); ch2.set_categories(cats)
ws.add_chart(ch2,"F"+str(sr-1))

# bunching
br=slast+3
ws.cell(row=br-1,column=1,value="Threshold gaming: solar projects bunch just under 50MW to stay in the local route").font=Font(name=FNT,bold=True,size=11,color=NAVY)
ws.cell(row=br,column=1,value="Capacity band"); ws.cell(row=br,column=2,value="Solar projects")
hdr(ws,br,2)
rr=br+1
for b,n in D["bunch"].items():
    ws.cell(row=rr,column=1,value=b); ws.cell(row=rr,column=2,value=n); ws.cell(row=rr,column=2).number_format="#,##0"; ws.cell(row=rr,column=1).font=Font(name=FNT,size=10)
    if b.startswith("45"):
        for c in (1,2): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=AMBER)
    rr+=1
ws.cell(row=rr+1,column=1,value="209 solar projects sit at 45-49.99MW vs just 5 at 50-55MW: developers size at 49.9MW to avoid the >50MW Nationally Significant Infrastructure Project (DCO) route.").font=Font(name=FNT,italic=True,size=9,color="595959")

order=['Read me','Overview','Funnel by technology','Funnel by nation','Durations','Time to decision','Volume vs application wtd','Decided vs awaiting','Solar timing explained','Decision route','Approval trend','Onshore wind by nation','Applications by tech','Decommissioned','Council control method','Outcomes by council party','Onshore wind by council party','Contested vs uncontested','Council vs national']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save('REPD_planning_outcomes_phase1.xlsx')
print("saved",len(wb.sheetnames),"sheets")
