import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

D=json.load(open('phase3b_land.json'))
NAVY="1F3864"; BLUE="2E5496"; AMBER="FFE699"; GREEN="C6E0B4"; FNT="Arial"
wb=load_workbook('REPD_planning_outcomes_phase1.xlsx')
if "Council land & MW density" in wb.sheetnames: del wb["Council land & MW density"]
ws=wb.create_sheet("Council land & MW density")
ws.sheet_view.showGridLines=False

def hdr(ws,row,n,fill=BLUE):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name=FNT,bold=True,color="FFFFFF",size=10); cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

ws["A1"]="Who approved it: capacity approved per km2 of land controlled"
ws["A1"].font=Font(name=FNT,bold=True,size=14,color=NAVY)
ws["A2"]="Each approval credited to the party controlling the council at the PERMIT date (planning permission granted), 2015-2026. Normalised by land-years controlled."
ws["A2"].font=Font(name=FNT,italic=True,size=9,color="595959")

cols=["Controlling party (at permit)","Approved (MW)","Projects","Share of approvals","Avg land controlled (km2)","Land share","Approved MW per km2 per yr"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
hdr(ws,r0,len(cols))
rr=r0+1
for o in D["rows"]:
    ws.cell(row=rr,column=1,value=o["party"]); ws.cell(row=rr,column=2,value=o["approvedMW"]); ws.cell(row=rr,column=3,value=o["nApproved"])
    ws.cell(row=rr,column=4,value=o["approvedSharePct"]/100); ws.cell(row=rr,column=5,value=o["avgLandKm2"]); ws.cell(row=rr,column=6,value=o["landSharePct"]/100)
    ws.cell(row=rr,column=7,value=o["mwPerKm2PerYr"])
    ws.cell(row=rr,column=4).number_format="0.0%"; ws.cell(row=rr,column=6).number_format="0.0%"; ws.cell(row=rr,column=7).number_format="0.000"
    for c in [2,3,5]: ws.cell(row=rr,column=c).number_format="#,##0"
    isni=o["party"] in ("Sinn Fein","DUP")
    ws.cell(row=rr,column=1).font=Font(name=FNT,size=10,bold=(o["party"] in("Con","Lab")),italic=isni)
    if o["party"]=="Con":
        for c in range(1,8): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=AMBER)
    if o["party"]=="Lab":
        for c in range(1,8): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=GREEN)
    rr+=1
last=rr-1
ws.cell(row=last+2,column=1,value="Conservative councils approved the most capacity (41,700 MW, 42%) - but on 42% of the land, that is just 0.035 MW/km2/yr. Labour approved 26,300 MW (27%) on only 11% of the land: 0.083 - about 2.4x denser. Every land-rich rural party (Lib Dem 0.044, SNP 0.019, independent-led 0.015) approves sparsely; urban Labour and Green (0.095) approve densely. The parties responsible for the most land deliver the least capacity from it.").font=Font(name=FNT,italic=True,size=9,color="595959")
ws.cell(row=last+2,column=1).alignment=Alignment(wrap_text=True)
ws.merge_cells(start_row=last+2,start_column=1,end_row=last+4,end_column=7)
ws.cell(row=last+6,column=1,value="CAVEATS: (1) Approval credited to control at the permit (planning-permission-granted) date - the 'who approved' question. Control at consent->construction->commissioning is a separate question (see Durations / build-out). (2) Reform's 0.179 sits on a tiny, brand-new base (71 projects, councils held only from 2025) - not comparable. (3) Excludes the national/strategic route, county-tier authorities, pre-reorganisation district approvals and approvals with no permit date. (4) Land-years for two-tier areas reorganised into unitaries (e.g. North Yorkshire, Cumbria, Somerset) are approximate. Direction and rough magnitude are robust; treat exact figures as indicative.").font=Font(name=FNT,size=8,italic=True,color="808080")
ws.cell(row=last+6,column=1).alignment=Alignment(wrap_text=True,vertical="top")
ws.merge_cells(start_row=last+6,start_column=1,end_row=last+10,end_column=7)
for col,w in {"A":24,"B":13,"C":10,"D":14,"E":18,"F":11,"G":20}.items(): ws.column_dimensions[col].width=w
ch=BarChart(); ch.type="col"; ch.title="Approved MW per km2 per year, by controlling party at permit"; ch.height=9; ch.width=18; ch.legend=None
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last); data=Reference(ws,min_col=7,min_row=r0,max_row=last)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
ws.add_chart(ch,"I4")

order=['Read me','Overview','Funnel by technology','Funnel by nation','Durations','Time to decision','Volume vs application wtd','Capacity vs count approval','Decided vs awaiting','Solar timing explained','Decision route','Approval trend','Onshore wind by nation','Applications by tech','Decommissioned','Council control method','Outcomes by council party','Onshore wind by council party','Contested vs uncontested','Council land & MW density','Council vs national','MP party method','Outcomes by MP party','Council vs MP']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save('REPD_planning_outcomes_phase1.xlsx')
print("saved",len(wb.sheetnames),"sheets")
