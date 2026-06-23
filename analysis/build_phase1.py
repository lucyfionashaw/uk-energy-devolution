import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

D = json.load(open('phase1_data.json'))

NAVY="1F3864"; BLUE="2E5496"; LIGHT="D9E1F2"; GREY="F2F2F2"; AMBER="FFE699"; RED="C00000"; GREEN="375623"
FNT="Arial"
wb = Workbook()

def style_header(ws, row, ncol, fill=BLUE, color="FFFFFF"):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FNT, bold=True, color=color, size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="BFBFBF"))

def title_block(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(name=FNT, bold=True, size=14, color=NAVY)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name=FNT, italic=True, size=9, color="595959")

def setw(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def body_font(ws, r0, r1, c0, c1):
    for r in range(r0, r1+1):
        for c in range(c0, c1+1):
            cell = ws.cell(row=r, column=c)
            if cell.font is None or cell.font.name != FNT:
                cell.font = Font(name=FNT, size=10)

# ---------------- README ----------------
ws = wb.active; ws.title = "Read me"
ws.sheet_view.showGridLines = False
title_block(ws, "UK Renewable Energy Planning Database — Planning Outcomes & Bottlenecks",
            "Phase 1 of the devolution analysis · all technologies · Great Britain + Northern Ireland")
notes = [
 ("Source", "Renewable Energy Planning Database (REPD), DESNZ, Q1 2026 extract (published April 2026). 14,316 project records."),
 ("URL", "https://www.gov.uk/government/publications/renewable-energy-planning-database-quarterly-extract"),
 ("Scope", "Every technology and every UK nation in the REPD. Capacity figures are Installed Capacity (MWelec)."),
 ("", ""),
 ("WHAT THIS WORKBOOK COVERS", "The planning 'stage-gate' journey: applied -> approved -> under construction -> operational, plus the ways projects die (refused, withdrawn, abandoned, expired) or stall (still in planning). Cut by technology, nation and year."),
 ("", ""),
 ("KEY DEFINITIONS", ""),
 ("Ever granted", "Project has a Planning Permission Granted date OR a current status of Awaiting Construction / Under Construction / Operational / Expired / Decommissioned. Counts anything that ever secured consent."),
 ("Ever built", "Has an Under Construction date OR status of Under Construction / Operational / Decommissioned."),
 ("Operational", "Has an Operational date OR status Operational / Decommissioned."),
 ("Refused", "Current status Application Refused, Appeal Refused or Secretary of State Refusal."),
 ("Withdrawn", "Sponsor withdrew the application (or the appeal)."),
 ("Abandoned", "Project recorded as abandoned by the developer."),
 ("Pending", "Still in the system: status Application Submitted or Appeal Lodged (i.e. no decision yet)."),
 ("Approval rate", "Ever granted / (ever granted + refused). Excludes withdrawn/abandoned/pending from the denominator."),
 ("National route", "Consenting authority is a national/strategic body (S36 / NSIP / Planning Inspectorate / Marine Scotland / MMO / devolved ministers) rather than a local council. Proxy for 'decided above the local level'."),
 ("", ""),
 ("CAVEATS", ""),
 ("Survivorship", "Median durations are for projects that REACHED a given gate. They understate true delay because stalled/dead projects never get an end date."),
 ("Snapshot", "REPD is a living register; statuses reflect the Q1 2026 extract and are periodically revised. Minor data-entry inconsistencies in status labels were normalised."),
 ("Approval-rate trend", "Post-2018 the apparent approval rate rises partly because hostile English onshore-wind applications stopped being submitted (see 'Onshore wind by nation')."),
 ("", ""),
 ("STATUS", "Phase 1 (REPD only). Phase 2 will join local-authority political control + MP party at decision date; Phase 3 will join the DESNZ/LCCC Contracts for Difference register."),
]
r = 4
for k, v in notes:
    ws.cell(row=r, column=1, value=k).font = Font(name=FNT, bold=True, size=10, color=NAVY)
    c = ws.cell(row=r, column=2, value=v); c.font = Font(name=FNT, size=10); c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
setw(ws, {"A": 22, "B": 110})

# ---------------- OVERVIEW ----------------
ws = wb.create_sheet("Overview")
ws.sheet_view.showGridLines = False
o = D["overall"]
title_block(ws, "Overview — the national picture", "Counts are project records; capacity is MW (GW = MW/1000).")
# headline funnel table
hdr = ["Stage / outcome", "Projects", "Capacity (MW)", "Capacity (GW)", "% of all projects"]
r0 = 4
ws.append([]) if False else None
for i, h in enumerate(hdr, 1):
    ws.cell(row=r0, column=i, value=h)
style_header(ws, r0, len(hdr))
rows = [
 ("Applied (all records)", o["n"], o["cap"]),
 ("Ever granted consent", o["granted"], o["grantedCap"]),
 ("Ever started construction", o["built"], o["builtCap"]),
 ("Operational", o["op"], o["opCap"]),
 ("Decommissioned", o["decom"], None),
 ("Refused", o["refused"], o["refusedCap"]),
 ("Withdrawn by sponsor", o["withdrawn"], o["withdrawnCap"]),
 ("Abandoned", o["abandoned"], None),
 ("Consent expired (lapsed)", o["expired"], None),
 ("Still pending a decision", o["pending"], o["pendingCap"]),
]
rr = r0+1
for label, n, cap in rows:
    ws.cell(row=rr, column=1, value=label)
    ws.cell(row=rr, column=2, value=n)
    if cap is not None:
        ws.cell(row=rr, column=3, value=cap)
        ws.cell(row=rr, column=4, value=f"=C{rr}/1000")
    ws.cell(row=rr, column=5, value=f"=B{rr}/$B${r0+1}")
    rr += 1
last = rr-1
for r in range(r0+1, last+1):
    ws.cell(row=r, column=2).number_format = "#,##0"
    ws.cell(row=r, column=3).number_format = "#,##0"
    ws.cell(row=r, column=4).number_format = "#,##0.0"
    ws.cell(row=r, column=5).number_format = "0.0%"
    ws.cell(row=r, column=1).font = Font(name=FNT, size=10, bold=(r==r0+1))
# highlight key rows
for r, fill in [(r0+1, LIGHT), (r0+10, AMBER)]:
    for c in range(1,6):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=fill)
# KPI callouts
k = last+3
kpis = [
 ("Headline approval rate", f"{o['approvalRatePct']:.1f}%", "of decided projects are granted — refusal is NOT the main blocker"),
 ("Consented but not yet built", f"{(o['grantedCap']-o['builtCap'])/1000:.0f} GW", f"{o['granted']-o['built']:,} projects hold consent but have not started construction"),
 ("Stuck in the queue", f"{o['pendingCap']/1000:.1f} GW", f"{o['pending']:,} projects awaiting a planning decision"),
 ("Routed above local level", f"{o['nationalPct']:.1f}%", "of all projects are consented nationally, not by a council"),
]
ws.cell(row=k, column=1, value="Key takeaways").font = Font(name=FNT, bold=True, size=11, color=NAVY)
k+=1
for label, val, desc in kpis:
    ws.cell(row=k, column=1, value=label).font = Font(name=FNT, bold=True, size=10)
    vc = ws.cell(row=k, column=2, value=val); vc.font = Font(name=FNT, bold=True, size=11, color=BLUE)
    ws.cell(row=k, column=3, value=desc).font = Font(name=FNT, size=10, italic=True, color="595959")
    k+=1
# pending aging mini-table
k+=1
ws.cell(row=k, column=1, value="How long pending projects have been waiting").font = Font(name=FNT, bold=True, size=11, color=NAVY)
k+=1
ws.cell(row=k, column=1, value="Wait band"); ws.cell(row=k, column=2, value="Projects"); ws.cell(row=k, column=3, value="Capacity (MW)")
style_header(ws, k, 3)
pa = D["pendingAging"]
order = ["<1 yr","1-2 yrs","2-3 yrs","3-5 yrs","5+ yrs"]
k+=1
for b in order:
    ws.cell(row=k, column=1, value=b); ws.cell(row=k, column=2, value=pa[b]["n"]); ws.cell(row=k, column=3, value=pa[b]["cap"])
    ws.cell(row=k, column=2).number_format="#,##0"; ws.cell(row=k, column=3).number_format="#,##0"
    if b in ("3-5 yrs","5+ yrs"):
        for c in range(1,4): ws.cell(row=k, column=c).fill=PatternFill("solid", fgColor=AMBER)
    k+=1
setw(ws, {"A":34,"B":14,"C":16,"D":14,"E":18})

# ---------------- FUNNEL BY TECH ----------------
def funnel_sheet(name, data, keycol, keyhdr):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    title_block(ws, name, "Counts = project records. Capacity = MW. Approval rate & built rate are Excel formulas.")
    cols = [keyhdr,"Projects","Capacity (MW)","Ever granted","Granted cap (MW)","Ever built","Built cap (MW)","Operational","Refused","Withdrawn","Abandoned","Expired","Pending","Pending cap (MW)","Decom.","National %","Approval rate","Built/consented"]
    r0=4
    for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
    style_header(ws, r0, len(cols))
    rr=r0+1
    for row in data:
        vals=[row[keycol],row["n"],row["cap"],row["granted"],row["grantedCap"],row["built"],row["builtCap"],row["op"],
              row["refused"],row["withdrawn"],row["abandoned"],row["expired"],row["pending"],row["pendingCap"],row["decom"],row["nationalPct"]/100]
        for i,v in enumerate(vals,1): ws.cell(row=rr,column=i,value=v)
        ws.cell(row=rr,column=17,value=f"=IF((D{rr}+I{rr})=0,\"\",D{rr}/(D{rr}+I{rr}))")  # approval = granted/(granted+refused)
        ws.cell(row=rr,column=18,value=f"=IF(D{rr}=0,\"\",F{rr}/D{rr})")  # built/consented
        rr+=1
    last=rr-1
    for r in range(r0+1,last+1):
        for c in [2,3,4,5,6,7,8,9,10,11,12,13,14,15]: ws.cell(row=r,column=c).number_format="#,##0"
        for c in [16,17,18]: ws.cell(row=r,column=c).number_format="0.0%"
        ws.cell(row=r,column=1).font=Font(name=FNT,size=10)
        if str(ws.cell(row=r,column=1).value) in ("Wind Onshore","Wind Offshore"):
            for c in range(1,19): ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor=GREY)
    ws.freeze_panes="B5"
    widths={"A":30}
    for c in range(2,19): widths[get_column_letter(c)]=12
    widths["Q"]=12; widths["R"]=14
    setw(ws, widths)
    return ws, r0, last

wsT, r0T, lastT = funnel_sheet("Funnel by technology", D["funnelTech"], "tech", "Technology")
wsC, r0C, lastC = funnel_sheet("Funnel by nation", D["funnelCountry"], "country", "Nation")

# bar chart: top techs consented vs built capacity (use first 6 rows = biggest)
chart = BarChart(); chart.type="col"; chart.title="Capacity by stage — top technologies (MW)"; chart.height=8; chart.width=18
cats = Reference(wsT, min_col=1, min_row=r0T+1, max_row=r0T+6)
for col,nm in [(3,"Applied"),(5,"Consented"),(7,"Built")]:
    data = Reference(wsT, min_col=col, min_row=r0T, max_row=r0T+6)
    chart.add_data(data, titles_from_data=True)
chart.set_categories(cats); chart.y_axis.title="MW"; chart.x_axis.title=None
wsT.add_chart(chart, "A30")

# ---------------- DURATIONS ----------------
ws = wb.create_sheet("Durations")
ws.sheet_view.showGridLines = False
title_block(ws, "Time through the pipeline (median months)", "Survivor-based: only projects that reached each gate. Years column = total app->operational.")
cols=["Technology","n (granted)","App -> Granted","Granted -> Construction","Construction -> Operational","App -> Operational (months)","App -> Operational (years)"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
style_header(ws, r0, len(cols))
rr=r0+1
for row in D["durations"]:
    ws.cell(row=rr,column=1,value=row["tech"])
    ws.cell(row=rr,column=2,value=row["nAppToGrant"])
    ws.cell(row=rr,column=3,value=row["medAppToGrant"])
    ws.cell(row=rr,column=4,value=row["medGrantToConstr"])
    ws.cell(row=rr,column=5,value=row["medConstrToOp"])
    ws.cell(row=rr,column=6,value=row["medAppToOp"])
    ws.cell(row=rr,column=7,value=f"=F{rr}/12")
    rr+=1
last=rr-1
for r in range(r0+1,last+1):
    ws.cell(row=r,column=2).number_format="#,##0"
    for c in [3,4,5,6]: ws.cell(row=r,column=c).number_format="0.0"
    ws.cell(row=r,column=7).number_format="0.0"
    ws.cell(row=r,column=1).font=Font(name=FNT,size=10)
    if str(ws.cell(row=r,column=1).value) in ("Wind Onshore","Wind Offshore","Solar Photovoltaics"):
        for c in range(1,8): ws.cell(row=r,column=c).fill=PatternFill("solid",fgColor=GREY)
setw(ws, {"A":32,"B":12,"C":16,"D":22,"E":24,"F":24,"G":22})
chart=BarChart(); chart.type="bar"; chart.title="Median months: application to consent"; chart.height=9; chart.width=16
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last)
data=Reference(ws,min_col=3,min_row=r0,max_row=last)
chart.add_data(data,titles_from_data=True); chart.set_categories(cats); chart.legend=None
ws.add_chart(chart,"I4")

# ---------------- DECISION ROUTE ----------------
ws = wb.create_sheet("Decision route")
ws.sheet_view.showGridLines = False
title_block(ws, "Local vs national consenting route", "National = S36 / NSIP / Planning Inspectorate / Marine Scotland / MMO / devolved ministers.")
ws.cell(row=4,column=1,value="By nation").font=Font(name=FNT,bold=True,size=11,color=NAVY)
cols=["Nation","Local","National","National %"]
r0=5
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
style_header(ws,r0,4)
rr=r0+1
for row in D["routeByNation"]:
    ws.cell(row=rr,column=1,value=row["country"]); ws.cell(row=rr,column=2,value=row["local"]); ws.cell(row=rr,column=3,value=row["national"])
    ws.cell(row=rr,column=4,value=f"=C{rr}/(B{rr}+C{rr})"); ws.cell(row=rr,column=4).number_format="0.0%"
    ws.cell(row=rr,column=2).number_format="#,##0"; ws.cell(row=rr,column=3).number_format="#,##0"
    rr+=1
natlast=rr-1
r2=rr+2
ws.cell(row=r2-1,column=1,value="By technology").font=Font(name=FNT,bold=True,size=11,color=NAVY)
for i,h in enumerate(["Technology","Local","National","National %"],1): ws.cell(row=r2,column=i,value=h)
style_header(ws,r2,4)
rr=r2+1
for row in D["routeByTech"]:
    ws.cell(row=rr,column=1,value=row["tech"]); ws.cell(row=rr,column=2,value=row["local"]); ws.cell(row=rr,column=3,value=row["national"])
    ws.cell(row=rr,column=4,value=f"=IF((B{rr}+C{rr})=0,\"\",C{rr}/(B{rr}+C{rr}))"); ws.cell(row=rr,column=4).number_format="0.0%"
    ws.cell(row=rr,column=2).number_format="#,##0"; ws.cell(row=rr,column=3).number_format="#,##0"
    rr+=1
setw(ws,{"A":32,"B":12,"C":12,"D":12})

# ---------------- APPROVAL TREND ----------------
ws = wb.create_sheet("Approval trend")
ws.sheet_view.showGridLines = False
title_block(ws,"Approval rate by year of decision","All technologies vs onshore wind. Note: post-2018 the all-tech rate rises as solar/battery volume dominates.")
cols=["Year","All: granted","All: refused","All approval %","Onshore granted","Onshore refused","Onshore approval %"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
style_header(ws,r0,len(cols))
ov={x["year"]:x for x in D["approvalOverallByYear"]}; owd={x["year"]:x for x in D["approvalOnshoreByYear"]}
years=sorted(ov)
rr=r0+1
for y in years:
    a=ov[y]; w=owd.get(y,{"granted":0,"refused":0})
    ws.cell(row=rr,column=1,value=str(y))
    ws.cell(row=rr,column=2,value=a["granted"]); ws.cell(row=rr,column=3,value=a["refused"])
    ws.cell(row=rr,column=4,value=f"=IF((B{rr}+C{rr})=0,\"\",B{rr}/(B{rr}+C{rr}))")
    ws.cell(row=rr,column=5,value=w["granted"]); ws.cell(row=rr,column=6,value=w["refused"])
    ws.cell(row=rr,column=7,value=f"=IF((E{rr}+F{rr})=0,\"\",E{rr}/(E{rr}+F{rr}))")
    for c in [4,7]: ws.cell(row=rr,column=c).number_format="0.0%"
    rr+=1
last=rr-1
setw(ws,{"A":8,"B":12,"C":12,"D":14,"E":14,"F":14,"G":16})
ch=LineChart(); ch.title="Approval rate: all tech vs onshore wind"; ch.height=9; ch.width=18; ch.y_axis.numFmt="0%"; ch.y_axis.title="Approval rate"
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last)
for col in [4,7]:
    data=Reference(ws,min_col=col,min_row=r0,max_row=last); ch.add_data(data,titles_from_data=True)
ch.set_categories(cats)
ws.add_chart(ch,"I4")

# ---------------- ONSHORE WIND BY NATION ----------------
ws = wb.create_sheet("Onshore wind by nation")
ws.sheet_view.showGridLines = False
title_block(ws,"Onshore wind applications submitted, by nation","The natural experiment: England's 2015 NPPF veto vs devolved Scotland/Wales/NI.")
cols=["Year","England","Scotland","Wales","Northern Ireland"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
style_header(ws,r0,len(cols))
rr=r0+1
for row in D["onshoreAppsByNation"]:
    ws.cell(row=rr,column=1,value=str(row["year"]))
    ws.cell(row=rr,column=2,value=row["England"]); ws.cell(row=rr,column=3,value=row["Scotland"])
    ws.cell(row=rr,column=4,value=row["Wales"]); ws.cell(row=rr,column=5,value=row["NI"])
    if row["year"] in (2015,2016):
        for c in range(1,6): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=AMBER)
    rr+=1
last=rr-1
ws.cell(row=last+2,column=1,value="2015: NPPF 'local veto' footnote introduced (June 2015). England applications collapse the following year while Scotland continues.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":8,"B":12,"C":12,"D":10,"E":16})
ch=LineChart(); ch.title="Onshore wind applications submitted per year, by nation"; ch.height=10; ch.width=20; ch.y_axis.title="Applications"
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last)
data=Reference(ws,min_col=2,min_row=r0,max_col=5,max_row=last)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
ws.add_chart(ch,"G4")

# ---------------- APPLICATIONS BY TECH ----------------
ws = wb.create_sheet("Applications by tech")
ws.sheet_view.showGridLines = False
title_block(ws,"Applications submitted per year, by technology","The shift from onshore wind (2008-15) to the solar + battery surge (2021-).")
cols=["Year","Solar","Onshore wind","Battery","Offshore wind"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
style_header(ws,r0,len(cols))
rr=r0+1
for row in D["appsByTech"]:
    ws.cell(row=rr,column=1,value=str(row["year"]))
    ws.cell(row=rr,column=2,value=row["Solar"]); ws.cell(row=rr,column=3,value=row["OnshoreWind"])
    ws.cell(row=rr,column=4,value=row["Battery"]); ws.cell(row=rr,column=5,value=row["OffshoreWind"])
    rr+=1
last=rr-1
setw(ws,{"A":8,"B":12,"C":14,"D":10,"E":14})
ch=LineChart(); ch.title="Applications submitted per year, by technology"; ch.height=10; ch.width=20; ch.y_axis.title="Applications"
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last)
data=Reference(ws,min_col=2,min_row=r0,max_col=5,max_row=last)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
ws.add_chart(ch,"G4")

# ---------------- DECOMMISSIONED ----------------
ws = wb.create_sheet("Decommissioned")
ws.sheet_view.showGridLines = False
title_block(ws,"Decommissioned projects","32 records in total (incl. several EMEC tidal/wave test devices); principal sites listed.")
cols=["REPD Ref","Site","Operator","Technology","Capacity (MW)","Nation","Planning authority"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
style_header(ws,r0,len(cols))
rr=r0+1
for row in D["decommissioned"]:
    ws.cell(row=rr,column=1,value=row["ref"]); ws.cell(row=rr,column=2,value=row["site"]); ws.cell(row=rr,column=3,value=row["operator"])
    ws.cell(row=rr,column=4,value=row["tech"]); ws.cell(row=rr,column=5,value=float(row["capMW"])); ws.cell(row=rr,column=6,value=row["country"]); ws.cell(row=rr,column=7,value=row["authority"])
    ws.cell(row=rr,column=5).number_format="#,##0.0"
    for c in range(1,8): ws.cell(row=rr,column=c).font=Font(name=FNT,size=10)
    rr+=1
ws.freeze_panes="A5"
setw(ws,{"A":10,"B":34,"C":34,"D":24,"E":14,"F":16,"G":26})

wb.save("REPD_planning_outcomes_phase1.xlsx")
print("saved")
