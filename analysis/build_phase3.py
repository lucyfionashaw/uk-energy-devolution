import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

D=json.load(open('phase3_land.json'))
NAVY="1F3864"; BLUE="2E5496"; AMBER="FFE699"; GREEN="C6E0B4"; RED="C00000"; FNT="Arial"
wb=load_workbook('REPD_planning_outcomes_phase1.xlsx')
for nm in ["Capacity vs count approval","Council land & MW density"]:
    if nm in wb.sheetnames: del wb[nm]

def hdr(ws,row,n,fill=BLUE):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name=FNT,bold=True,color="FFFFFF",size=10); cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def title(ws,t,s=None):
    ws["A1"]=t; ws["A1"].font=Font(name=FNT,bold=True,size=14,color=NAVY)
    if s: ws["A2"]=s; ws["A2"].font=Font(name=FNT,italic=True,size=9,color="595959")
def setw(ws,w):
    for k,v in w.items(): ws.column_dimensions[k].width=v

# ---------- APPROVAL: the denominator matters ----------
ws=wb.create_sheet("Capacity vs count approval"); ws.sheet_view.showGridLines=False
title(ws,"What is the approval rate? The denominator is everything","Conditioning on a decision (90%) hides the real blocker: councils not deciding. The honest headline is approved / applied.")
co=D["cohort"]
ws.cell(row=4,column=1,value="The headline that matters: approved capacity / ALL applied capacity").font=Font(name=FNT,bold=True,size=11,color=NAVY)
cols=["Cohort (by application date)","Applied (GW)","Approved so far (GW)","Approved %","Still pending (GW)"]
r0=5
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
hdr(ws,r0,len(cols))
rows=[("All-time",co["allTime"]["appliedGW"],co["allTime"]["approvedGW"],co["allTime"]["approvedPct"],None),
      ("Applied since 2019 (current data)",co["since2019"]["appliedGW"],co["since2019"]["approvedGW"],co["since2019"]["approvedPct"],co["since2019"]["pendingGW"]),
      ("Applied in last 5 yrs (2021+)",co["since2021"]["appliedGW"],co["since2021"]["approvedGW"],co["since2021"]["approvedPct"],co["since2021"]["pendingGW"]),
      ("Same 2019+ cohort, as at mid-2024 (prior analysis)",co["prior2024paper"]["appliedGW"],co["prior2024paper"]["approvedGW"],co["prior2024paper"]["approvedPct"],None)]
rr=r0+1
for lab,ap,apr,pct,pend in rows:
    ws.cell(row=rr,column=1,value=lab); ws.cell(row=rr,column=2,value=ap); ws.cell(row=rr,column=3,value=apr)
    ws.cell(row=rr,column=4,value=pct/100); ws.cell(row=rr,column=4).number_format="0.0%"
    if pend is not None: ws.cell(row=rr,column=5,value=pend)
    for c in [2,3,5]: ws.cell(row=rr,column=c).number_format="#,##0.0"
    if "mid-2024" in lab:
        for c in range(1,6): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=AMBER)
    rr+=1
ws.cell(row=rr+1,column=1,value="Only ~42-58% of applied capacity has been approved - and the rate FALLS as the pipeline grows, because so much capacity is stuck undecided (the 'planning purgatory' pile, ~88 GW). A project waiting years for a decision is effectively blocked, even though it never appears as a refusal.").font=Font(name=FNT,size=10,color=RED); ws.cell(row=rr+1,column=1).alignment=Alignment(wrap_text=True)
ws.merge_cells(start_row=rr+1,start_column=1,end_row=rr+2,end_column=6)
ws.cell(row=rr+4,column=1,value="Why the '90%' figure is misleading on its own").font=Font(name=FNT,bold=True,size=11,color=NAVY)
ws.cell(row=rr+5,column=1,value="'90% approved' = granted / (granted + refused). It conditions on a decision having been made, so it ignores the biggest failure mode - non-decision / delay. It is a true but narrow sub-statistic, not the planning success rate. (It is ~90% whether weighted by project count or by capacity - large projects are not refused more, just delayed more.)").font=Font(name=FNT,size=10); ws.cell(row=rr+5,column=1).alignment=Alignment(wrap_text=True)
ws.merge_cells(start_row=rr+5,start_column=1,end_row=rr+6,end_column=6)
gr=rr+8
ws.cell(row=gr-1,column=1,value="And of what IS approved, very little is built").font=Font(name=FNT,bold=True,size=11,color=NAVY)
ws.cell(row=gr,column=1,value="Of capacity approved since 2019, share now operational"); ws.cell(row=gr,column=2,value=co["since2019"]["operationalPctOfApproved"]/100); ws.cell(row=gr,column=2).number_format="0.0%"
ws.cell(row=gr+1,column=1,value="Of all approved capacity (all-time), share now operational"); ws.cell(row=gr+1,column=2,value=co["allTime"]["operationalPctOfApproved"]/100); ws.cell(row=gr+1,column=2).number_format="0.0%"
ws.cell(row=gr,column=1).font=Font(name=FNT,size=10); ws.cell(row=gr+1,column=1).font=Font(name=FNT,size=10)
setw(ws,{"A":52,"B":14,"C":18,"D":12,"E":16,"F":6})

# ---------- LAND AREA & MW DENSITY ----------
ws=wb.create_sheet("Council land & MW density"); ws.sheet_view.showGridLines=False
title(ws,"Land area controlled vs renewables approved per km2","The land-rich parties build the least per acre. Current (2026) council control x approved capacity to date.")
cols=["Controlling party","Councils","Land controlled (km2)","Land share","Approved capacity (MW)","Approved share","Approved MW per km2","Applied MW per km2"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
hdr(ws,r0,len(cols))
rr=r0+1
for o in D["landRows"]:
    ws.cell(row=rr,column=1,value=o["party"]); ws.cell(row=rr,column=2,value=o["councils"]); ws.cell(row=rr,column=3,value=o["landKm2"])
    ws.cell(row=rr,column=4,value=o["landSharePct"]/100); ws.cell(row=rr,column=5,value=o["approvedMW"]); ws.cell(row=rr,column=6,value=o["approvedSharePct"]/100)
    ws.cell(row=rr,column=7,value=o["mwPerKm2"]); ws.cell(row=rr,column=8,value=o["appliedMWperKm2"])
    ws.cell(row=rr,column=4).number_format="0.0%"; ws.cell(row=rr,column=6).number_format="0.0%"
    ws.cell(row=rr,column=7).number_format="0.000"; ws.cell(row=rr,column=8).number_format="0.000"
    for c in [2,3,5]: ws.cell(row=rr,column=c).number_format="#,##0"
    isni=o["party"] in ("Sinn Fein","DUP")
    ws.cell(row=rr,column=1).font=Font(name=FNT,size=10,bold=(o["party"] in("Con","Lab")),italic=isni)
    if o["party"] in ("Con","Lab"):
        for c in range(1,9): ws.cell(row=rr,column=c).fill=PatternFill("solid",fgColor=(AMBER if o["party"]=="Con" else GREEN))
    rr+=1
last=rr-1
ws.cell(row=last+2,column=1,value="Conservative councils control ~60,000 km2 (3x Labour's 21,000) but approve just 0.68 MW/km2 vs Labour's 1.13 - Labour squeezes ~1.65x more renewable capacity from each km2 it controls. The rural Lib Dem (0.44), SNP (0.32) and independent-led (0.27) councils hold most of GB's land yet are the least dense.").font=Font(name=FNT,italic=True,size=9,color="595959"); ws.cell(row=last+2,column=1).alignment=Alignment(wrap_text=True)
ws.merge_cells(start_row=last+2,start_column=1,end_row=last+3,end_column=8)
ws.cell(row=last+5,column=1,value="CAVEATS: (1) Snapshot - current control x cumulative approvals; it conflates WHERE capacity sits with WHO approved it (control changes hands). Reform's high density especially reflects 2025 gains of councils with pre-existing pipelines, not Reform approvals. (2) Land = ONS Standard Area Measurement land area (Dec 2024), lower-tier councils; county-tier and a few abolished-district projects (8 councils) are excluded. (3) Density reflects both land type (urban Labour hosts dense battery/solar; rural shires host contested onshore wind) and approval behaviour.").font=Font(name=FNT,size=8,italic=True,color="808080"); ws.cell(row=last+5,column=1).alignment=Alignment(wrap_text=True,vertical="top")
ws.merge_cells(start_row=last+5,start_column=1,end_row=last+8,end_column=8)
setw(ws,{"A":18,"B":9,"C":18,"D":10,"E":18,"F":12,"G":16,"H":16})
ch=BarChart(); ch.type="col"; ch.title="Approved MW per km2 of land controlled, by party"; ch.height=9; ch.width=17; ch.legend=None
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last); data=Reference(ws,min_col=7,min_row=r0,max_row=last)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
ws.add_chart(ch,"J4")

order=['Read me','Overview','Funnel by technology','Funnel by nation','Durations','Time to decision','Volume vs application wtd','Capacity vs count approval','Decided vs awaiting','Solar timing explained','Decision route','Approval trend','Onshore wind by nation','Applications by tech','Decommissioned','Council control method','Outcomes by council party','Onshore wind by council party','Contested vs uncontested','Council land & MW density','Council vs national','MP party method','Outcomes by MP party','Council vs MP']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save('REPD_planning_outcomes_phase1.xlsx')
print("saved",len(wb.sheetnames),"sheets")
