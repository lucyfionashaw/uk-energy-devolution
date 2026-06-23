import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

D=json.load(open('phase2_data.json'))
NAVY="1F3864"; BLUE="2E5496"; GREY="F2F2F2"; AMBER="FFE699"; GREEN="C6E0B4"; FNT="Arial"
wb=load_workbook('REPD_planning_outcomes_phase1.xlsx')

def hdr(ws,row,n,fill=BLUE):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.font=Font(name=FNT,bold=True,color="FFFFFF",size=10)
        cell.fill=PatternFill("solid",fgColor=fill)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
def title(ws,t,s=None):
    ws["A1"]=t; ws["A1"].font=Font(name=FNT,bold=True,size=14,color=NAVY)
    if s: ws["A2"]=s; ws["A2"].font=Font(name=FNT,italic=True,size=9,color="595959")
def setw(ws,w):
    for k,v in w.items(): ws.column_dimensions[k].width=v

ORDER=["Con","Lab","LibDem","SNP","Plaid","Green","Reform","Other/Ind","Nationalist (pre-2007)"]
def partytable(ws,r0,data,label):
    cols=[label,"Projects","Granted","Refused","Approval rate","Capacity consented (MW)","Built (MW)"]
    for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
    hdr(ws,r0,len(cols))
    rr=r0+1
    for p in ORDER:
        if p not in data: continue
        o=data[p]
        ws.cell(row=rr,column=1,value=p); ws.cell(row=rr,column=2,value=o["n"]); ws.cell(row=rr,column=3,value=o["granted"]); ws.cell(row=rr,column=4,value=o["refused"])
        ws.cell(row=rr,column=5,value=f"=IF((C{rr}+D{rr})=0,\"\",C{rr}/(C{rr}+D{rr}))")
        ws.cell(row=rr,column=6,value=o.get("grantedCap","")); ws.cell(row=rr,column=7,value=o.get("builtCap",""))
        ws.cell(row=rr,column=5).number_format="0.0%"
        for c in [2,3,4,6,7]:
            if ws.cell(row=rr,column=c).value!="": ws.cell(row=rr,column=c).number_format="#,##0"
        ws.cell(row=rr,column=1).font=Font(name=FNT,size=10,bold=(p in("Con","Lab")))
        rr+=1
    return rr-1

# ---- METHOD / coverage ----
ws=wb.create_sheet("Council control method"); ws.sheet_view.showGridLines=False
title(ws,"Phase 2 — matching projects to council political control","Who controlled the council when each decision was made.")
cov=D["coverage"]
notes=[
 ("Council control source","Open Council Data UK (opencouncildata.co.uk), 'GB Compositions 1973-2026', CC0 public domain. Controlling party = largest party by seats each year; collapsed to change-points and matched to each project's planning authority."),
 ("National government","Date lookup: Con (to May 1997), Lab (1997-2010), Con-led Coalition (2010-15), Con (2015 - Jul 2024), Lab (Jul 2024-)."),
 ("Match date","Control taken at the DECISION year (grant/refusal); consent and commissioning dates also available for later passes."),
 ("",""),
 ("COVERAGE",""),
 ("Total project records",cov["total"]),
 ("National/strategic route (no council)",cov["national"]),
 ("Local-route projects",cov["localRoute"]),
 ("Matched to a council",cov["matched"]),
 ("Unmatched",cov["unmatched"]),
 ("",""),
 ("Match rate (of GB local route)","95.6% (12,805 of 13,397)"),
 ("",""),
 ("CAVEATS",""),
 ("Northern Ireland excluded","Most unmatched records are NI councils (Fermanagh & Omagh, Derry, etc.) - GB composition file does not cover NI. NI uses a separate, shorter (2014-) series."),
 ("Largest-party proxy","Control = largest party; coalitions/NOC are attributed to the largest single party. The underlying seat counts are in the source if finer control is needed."),
 ("Alignment confound","'Council aligned with national government' overlaps heavily with the Con-Con 2015-23 onshore-wind moratorium era - treat the alignment cut as descriptive, not causal."),
 ("Pre-2007 'Nationalist'","Older Scottish/Welsh records combine SNP+Plaid as 'nat'; shown separately and not merged with modern SNP/Plaid."),
]
r=4
for k,v in notes:
    ws.cell(row=r,column=1,value=k).font=Font(name=FNT,bold=True,size=10,color=NAVY)
    c=ws.cell(row=r,column=2,value=v); c.font=Font(name=FNT,size=10); c.alignment=Alignment(wrap_text=True,vertical="top")
    if isinstance(v,int): ws.cell(row=r,column=2).number_format="#,##0"
    r+=1
setw(ws,{"A":34,"B":92})

# ---- OUTCOMES BY COUNCIL PARTY ----
ws=wb.create_sheet("Outcomes by council party"); ws.sheet_view.showGridLines=False
title(ws,"Planning outcomes by controlling council party","All technologies. Approval rate = granted / (granted + refused), control at decision year.")
last=partytable(ws,4,D["byParty"],"Council controlling party")
ws.cell(row=last+2,column=1,value="Across all technologies the gap is modest (Con 89% - SNP 93%): in aggregate, council political control is a weak predictor of approval.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":24,"B":11,"C":11,"D":11,"E":14,"F":22,"G":14})
ch=BarChart(); ch.type="bar"; ch.title="Approval rate by council party (all technologies)"; ch.height=8; ch.width=15; ch.y_axis.numFmt="0%"
cats=Reference(ws,min_col=1,min_row=5,max_row=last); data=Reference(ws,min_col=5,min_row=4,max_row=last)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.legend=None
ws.add_chart(ch,"I4")

# ---- ONSHORE WIND BY PARTY (the key finding) ----
ws=wb.create_sheet("Onshore wind by council party"); ws.sheet_view.showGridLines=False
title(ws,"Onshore wind: approval by controlling council party","The contested technology - where local political control actually bites.")
last=partytable(ws,4,D["byPartyOnshore"],"Council controlling party")
# highlight Con row
for r in range(5,last+1):
    if ws.cell(r,1).value=="Con":
        for c in range(1,8): ws.cell(r,c).fill=PatternFill("solid",fgColor=AMBER)
ws.cell(row=last+2,column=1,value="Conservative-controlled councils approve onshore wind at 61% vs Labour 69%, Lib Dem 71%, SNP 75%. Local control matters most for visible, contested technology.").font=Font(name=FNT,italic=True,size=9,color="595959")
ws.cell(row=last+3,column=1,value="By contrast, solar and battery approval exceeds 88% under every party (see comparison sheet) - uncontested tech is approved regardless of who controls the council.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":24,"B":11,"C":11,"D":11,"E":14,"F":22,"G":14})
ch=BarChart(); ch.type="bar"; ch.title="Onshore wind approval rate by council party"; ch.height=8; ch.width=15; ch.y_axis.numFmt="0%"
cats=Reference(ws,min_col=1,min_row=5,max_row=last); data=Reference(ws,min_col=5,min_row=4,max_row=last)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.legend=None
ws.add_chart(ch,"I4")

# ---- TECH COMPARISON: contested vs uncontested ----
ws=wb.create_sheet("Contested vs uncontested"); ws.sheet_view.showGridLines=False
title(ws,"Does council party matter? Onshore wind vs solar vs battery","Approval rate by controlling party, three technologies side by side.")
cols=["Council party","Onshore wind","Solar PV","Battery"]
r0=4
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
hdr(ws,r0,len(cols))
rr=r0+1
parties=["Con","Lab","LibDem","SNP","Other/Ind"]
for p in parties:
    ws.cell(row=rr,column=1,value=p)
    for ci,dset in enumerate([D["byPartyOnshore"],D["byPartySolar"],D["byPartyBattery"]],2):
        o=dset.get(p)
        ws.cell(row=rr,column=ci,value=(o["approvalRate"]/100 if o else None))
        ws.cell(row=rr,column=ci).number_format="0.0%"
    ws.cell(row=rr,column=1).font=Font(name=FNT,size=10)
    rr+=1
last=rr-1
ws.cell(row=last+2,column=1,value="Spread across parties: onshore wind ~14 pts (61-75%); solar ~5 pts (92-98%); battery ~6 pts (88-93%). Partisanship shows up only on the contested technology.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":16,"B":14,"C":12,"D":12})
ch=BarChart(); ch.type="col"; ch.title="Approval rate by council party and technology"; ch.height=9; ch.width=16; ch.y_axis.numFmt="0%"
cats=Reference(ws,min_col=1,min_row=r0+1,max_row=last)
for col in [2,3,4]:
    data=Reference(ws,min_col=col,min_row=r0,max_row=last); ch.add_data(data,titles_from_data=True)
ch.set_categories(cats)
ws.add_chart(ch,"F4")

# ---- ALIGNMENT ----
ws=wb.create_sheet("Council vs national"); ws.sheet_view.showGridLines=False
title(ws,"Does it matter if the council matches Westminster?","Approval where controlling council party = national government party, vs not.")
cols=["","Projects","Granted","Refused","Approval rate"]
r0=4
ws.cell(row=r0-1,column=1,value="All technologies").font=Font(name=FNT,bold=True,size=11,color=NAVY)
for i,h in enumerate(cols,1): ws.cell(row=r0,column=i,value=h)
hdr(ws,r0,len(cols))
def alignrows(ws,r0,data):
    rr=r0+1
    for lab,key in [("Council aligned with Westminster","aligned"),("Council differs from Westminster","misaligned")]:
        o=data[key]; ws.cell(row=rr,column=1,value=lab); ws.cell(row=rr,column=2,value=o["n"]); ws.cell(row=rr,column=3,value=o["granted"]); ws.cell(row=rr,column=4,value=o["refused"])
        ws.cell(row=rr,column=5,value=f"=C{rr}/(C{rr}+D{rr})"); ws.cell(row=rr,column=5).number_format="0.0%"
        for c in [2,3,4]: ws.cell(row=rr,column=c).number_format="#,##0"
        rr+=1
    return rr-1
last=alignrows(ws,r0,D["align"])
r2=last+3
ws.cell(row=r2-1,column=1,value="Onshore wind only").font=Font(name=FNT,bold=True,size=11,color=NAVY)
for i,h in enumerate(cols,1): ws.cell(row=r2,column=i,value=h)
hdr(ws,r2,len(cols))
last2=alignrows(ws,r2,D["alignOnshore"])
ws.cell(row=last2+2,column=1,value="All tech: alignment makes almost no difference (90.0% vs 88.8%). Onshore wind: 'aligned' is LOWER (60.9% vs 67.1%) - but this is mostly the Con-council/Con-government moratorium years, so read as descriptive, not causal.").font=Font(name=FNT,italic=True,size=9,color="595959")
setw(ws,{"A":34,"B":11,"C":11,"D":11,"E":14})

# reorder
order=['Read me','Overview','Funnel by technology','Funnel by nation','Durations','Time to decision','Decided vs awaiting','Solar timing explained','Decision route','Approval trend','Onshore wind by nation','Applications by tech','Decommissioned','Council control method','Outcomes by council party','Onshore wind by council party','Contested vs uncontested','Council vs national']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
wb.save('REPD_planning_outcomes_phase1.xlsx')
print("saved", len(wb.sheetnames),"sheets")
